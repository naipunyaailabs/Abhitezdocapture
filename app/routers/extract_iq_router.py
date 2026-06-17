from fastapi import APIRouter, UploadFile, File, Form, HTTPException, Depends
from fastapi.responses import Response
import time
import io
import json
import openpyxl
from app.services.extract_iq.extract_iq_service import extract_iq_service
from app.services.subscription_service import subscription_service
from app.services.history_service import history_service
from app.utils.auth import get_current_user, require_service

# Per-user access guard for this service (admins bypass; None = all allowed).
require_svc = require_service("extract-iq")
from app.models.user import UserResponse

router = APIRouter()


@router.post("/extract")
async def extract_iq_process(
    document: UploadFile = File(...),
    fields: str = Form(...),
    current_user: UserResponse = Depends(require_svc)
):
    """
    Extract handwritten & printed data from documents using user-defined fields.
    
    - document: PDF or image file
    - fields: JSON string of field definitions, e.g.:
      [
        {"key": "name", "label": "Full Name", "description": "Person's full name"},
        {"key": "date", "label": "Date", "description": "Date on the document"},
        ...
      ]
    """
    try:
        # Check subscription
        can_process, sub, message = await subscription_service.can_process(current_user.userId)
        if not can_process:
            raise HTTPException(
                status_code=403,
                detail=f"Processing limit reached. {message}. Please upgrade your plan."
            )
        
        # Parse field definitions
        try:
            field_definitions = json.loads(fields)
            if not isinstance(field_definitions, list) or len(field_definitions) == 0:
                raise ValueError("Fields must be a non-empty array")
        except (json.JSONDecodeError, ValueError) as e:
            raise HTTPException(status_code=400, detail=f"Invalid field definitions: {str(e)}")

        # Validate field definitions
        for i, field in enumerate(field_definitions):
            if not field.get("key"):
                raise HTTPException(
                    status_code=400,
                    detail=f"Field at index {i} is missing 'key' property"
                )

        start_time = time.time()
        buffer = await document.read()
        file_name = document.filename

        result = await extract_iq_service.extract_multi_page(
            buffer, file_name, field_definitions, current_user.userId
        )

        processing_time = int((time.time() - start_time) * 1000)

        # Record history
        await history_service.create_record({
            "userId": current_user.userId,
            "serviceId": "extract-iq",
            "serviceName": "ExtractIQ",
            "fileName": file_name,
            "fileSize": len(buffer),
            "format": "json",
            "status": "success",
            "result": f"Extracted {len(field_definitions)} fields across {result['total_pages']} pages",
            "processingTime": processing_time
        })

        # Increment usage
        await subscription_service.increment_usage(current_user.userId)

        return {
            "success": True,
            "data": result
        }
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ExtractIQRouter] Extract Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/export")
async def export_extract_iq(
    data: dict,
    current_user: UserResponse = Depends(require_svc)
):
    """
    Exports all validated records to Excel.
    Each page/record becomes one row in the spreadsheet.

    Expects data format:
    {
      "validated_records": [
        {
          "page_number": 1,
          "fields": {
            "field_key": {"value": "...", "edited_value": "..."},
            ...
          }
        }
      ],
      "field_definitions": [
        {"key": "field_key", "label": "Field Label"},
        ...
      ]
    }
    """
    try:
        records = data.get("validated_records", [])
        field_definitions = data.get("field_definitions", [])

        if not records:
            raise HTTPException(status_code=400, detail="No records to export")

        # Create new Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ExtractIQ Export"

        # Build field key to label mapping from definitions
        field_to_header = {}
        ordered_keys = []
        for field in field_definitions:
            key = field.get("key", "")
            label = field.get("label", key.replace("_", " ").title())
            if key:
                field_to_header[key] = label
                ordered_keys.append(key)

        # If no field definitions provided, infer from records
        if not ordered_keys:
            all_keys = set()
            for record in records:
                all_keys.update(record.get("fields", {}).keys())
            ordered_keys = sorted(all_keys)
            field_to_header = {k: k.replace("_", " ").title() for k in ordered_keys}

        # Create headers row
        headers = [field_to_header.get(k, k.replace("_", " ").title()) for k in ordered_keys]
        ws.append(headers)

        # Add data rows
        for record in records:
            fields = record.get("fields", {})
            row = []
            for field_key in ordered_keys:
                field_obj = fields.get(field_key, {})
                # Use edited_value if present (user modified), else original value
                val = field_obj.get("edited_value")
                if val is None:
                    val = field_obj.get("value", "N/A")
                if val is None or val == "":
                    val = "N/A"
                row.append(val)
            ws.append(row)

        # Style header row - ExtractIQ uses an amber/yellow theme
        header_fill = openpyxl.styles.PatternFill(
            start_color="FBBF24", end_color="FBBF24", fill_type="solid"
        )
        header_font = openpyxl.styles.Font(bold=True, color="000000")

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"ExtractIQ_{len(records)}_Records_{int(time.time())}.xlsx"
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        print(f"[ExtractIQRouter] Export Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
