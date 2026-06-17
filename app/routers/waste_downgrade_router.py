"""
Waste & Downgrade Router — API endpoints for the
"Wastage + Downgrade" sheet extractor.

Endpoints:
  POST /api/waste-downgrade/extract  → Extract a single document
  POST /api/waste-downgrade/export   → Export the extracted data as Excel
"""

from __future__ import annotations

import io
import time
import json
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel

from app.services.waste_downgrade.waste_downgrade_service import (
    COLUMNS,
    waste_downgrade_service,
)
from app.services.subscription_service import subscription_service
from app.services.history_service import history_service
from app.utils.auth import get_current_user, require_service

# Per-user access guard for this service (admins bypass; None = all allowed).
require_svc = require_service("waste-downgrade")
from app.models.user import UserResponse


router = APIRouter()

ALLOWED_EXTENSIONS = {
    ".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".webp",
}
MAX_FILE_SIZE_MB = 20


def _validate_file(filename: str, size: int):
    if not filename:
        raise HTTPException(400, "No filename provided.")
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(415, f"Unsupported file type '{ext}'.")
    if size > MAX_FILE_SIZE_MB * 1024 * 1024:
        raise HTTPException(413, f"File too large. Max: {MAX_FILE_SIZE_MB} MB.")


class ExportRequest(BaseModel):
    pages: List[Dict[str, Any]] = []
    title: str = "Waste_Downgrade_Export"


@router.post("/extract")
async def waste_downgrade_extract(
    document: UploadFile = File(...),
    current_user: UserResponse = Depends(require_svc),
):
    buffer = await document.read()
    _validate_file(document.filename, len(buffer))

    can_process, _sub, message = await subscription_service.can_process(current_user.userId)
    if not can_process:
        raise HTTPException(
            status_code=403,
            detail=f"Processing limit reached. {message}. Please upgrade your plan.",
        )

    start = time.time()
    try:
        result = await waste_downgrade_service.extract(buffer, document.filename)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        print(f"[WasteDowngradeRouter] Extraction error: {e}")
        raise HTTPException(500, str(e))

    processing_time = int((time.time() - start) * 1000)

    await subscription_service.increment_usage(current_user.userId)

    await history_service.create_record({
        "userId": current_user.userId,
        "serviceId": "waste-downgrade",
        "serviceName": "Waste & Downgrade",
        "fileName": document.filename,
        "fileSize": len(buffer),
        "format": "excel",
        "status": "success",
        "result": f"Extracted {result['total_rows']} rows across {result['total_pages']} pages",
        "processingTime": processing_time,
    })

    return {"success": True, "data": result}


@router.post("/extract-streaming")
async def waste_downgrade_extract_streaming(
    document: UploadFile = File(...),
    current_user: UserResponse = Depends(require_svc),
):
    """
    Extract and stream results as each page completes.
    Returns JSONL (JSON Lines) with progressive results.
    """
    buffer = await document.read()
    _validate_file(document.filename, len(buffer))

    can_process, _sub, message = await subscription_service.can_process(current_user.userId)
    if not can_process:
        raise HTTPException(
            status_code=403,
            detail=f"Processing limit reached. {message}. Please upgrade your plan.",
        )

    start = time.time()

    async def stream_results():
        try:
            async for chunk in waste_downgrade_service.extract_streaming(buffer, document.filename):
                yield chunk
        except ValueError as e:
            yield json.dumps({"type": "error", "error": str(e)}).encode() + b"\n"
        except Exception as e:
            print(f"[WasteDowngradeRouter] Streaming extraction error: {e}")
            import traceback
            traceback.print_exc()
            yield json.dumps({"type": "error", "error": str(e)}).encode() + b"\n"
        finally:
            processing_time = int((time.time() - start) * 1000)
            await subscription_service.increment_usage(current_user.userId)
            await history_service.create_record({
                "userId": current_user.userId,
                "serviceId": "waste-downgrade",
                "serviceName": "Waste & Downgrade",
                "fileName": document.filename,
                "fileSize": len(buffer),
                "format": "excel",
                "status": "success",
                "result": "Streaming extraction completed",
                "processingTime": processing_time,
            })

    return StreamingResponse(
        stream_results(),
        media_type="application/x-ndjson",
        headers={"X-Content-Type-Options": "nosniff"},
    )


@router.post("/export")
async def waste_downgrade_export(
    body: ExportRequest,
    current_user: UserResponse = Depends(require_svc),
):
    """Export extracted data preserving the source sheet layout:
        Row 1: DATE | <page date>
        Row 2: <merged> Wastage + Downgrade <merged>
        Row 3: IO NO | A-GRADE | ... | Total
        Rows 4+: data
    One sheet per page if multiple pages have distinct dates; otherwise a single sheet.
    """
    if not body.pages:
        raise HTTPException(400, "pages payload is required.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    super_header_fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    center = Alignment(horizontal="center", vertical="center")

    for idx, page in enumerate(body.pages, start=1):
        date_value = page.get("date", "") or ""
        rows = page.get("rows", []) or []

        title = f"Page {idx}"
        if len(title) > 31:
            title = title[:31]
        ws = wb.create_sheet(title=title)

        # Row 1: DATE label + value
        ws.cell(row=1, column=1, value="DATE").font = header_font
        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=1, column=2, value=date_value)

        n_cols = len(COLUMNS)
        # Row 2: super-header "Wastage + Downgrade" across grade/waste columns
        # Columns B..I cover A-GRADE..PATTI (indexes 2..9). Total stays on its own.
        ws.cell(row=2, column=2, value="Wastage + Downgrade").font = header_font
        ws.cell(row=2, column=2).fill = super_header_fill
        ws.cell(row=2, column=2).alignment = center
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=n_cols - 1)

        # Row 3: column headers
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # Rows 4+: data
        for r_offset, row in enumerate(rows, start=4):
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                ws.cell(row=r_offset, column=col_idx, value=str(row.get(col_name, "") or ""))

        # Auto column widths
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                if col_letter is None and hasattr(cell, "column_letter"):
                    col_letter = cell.column_letter
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 2, 24)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{body.title}_{int(time.time())}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
