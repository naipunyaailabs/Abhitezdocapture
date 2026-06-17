"""
Production Sheets Router — API endpoints for the Abhitex daily production
register extractor (reduced "green column" set).

Endpoints:
  POST /api/production-sheets/extract            → Extract a single document
  POST /api/production-sheets/extract-streaming  → Extract and stream per page
  POST /api/production-sheets/export             → Export extracted data as Excel
"""

from __future__ import annotations

import io
import json
import time
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel

from app.services.production_sheets.production_sheets_service import (
    COLUMNS,
    sheet_meta,
    production_sheets_service,
)
from app.services.subscription_service import subscription_service
from app.services.history_service import history_service
from app.utils.auth import get_current_user, require_service
from app.models.user import UserResponse

# Per-user access guard for this service (admins bypass; None = all allowed).
require_ps = require_service("production-sheets")


router = APIRouter()

ALLOWED_EXTENSIONS = {
    ".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".webp",
}
MAX_FILE_SIZE_MB = 20

SERVICE_ID = "production-sheets"
SERVICE_NAME = "Production Sheets"


def _validate_file(filename: str, size: int):
    if not filename:
        raise HTTPException(400, "No filename provided.")
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(415, f"Unsupported file type '{ext}'.")
    if size > MAX_FILE_SIZE_MB * 1024 * 1024:
        raise HTTPException(413, f"File too large. Max: {MAX_FILE_SIZE_MB} MB.")


class ExportRequest(BaseModel):
    pages: List[Dict[str, Any]] = []
    title: str = "Production_Sheets_Export"


@router.post("/extract")
async def production_sheets_extract(
    document: UploadFile = File(...),
    current_user: UserResponse = Depends(require_ps),
):
    buffer = await document.read()
    _validate_file(document.filename, len(buffer))

    can_process, _sub, message = await subscription_service.can_process(current_user.userId)
    if not can_process:
        raise HTTPException(
            status_code=403,
            detail=f"Processing limit reached. {message}. Please upgrade your plan.",
        )

    start = time.time()
    try:
        result = await production_sheets_service.extract(buffer, document.filename)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        print(f"[ProductionSheetsRouter] Extraction error: {e}")
        raise HTTPException(500, str(e))

    processing_time = int((time.time() - start) * 1000)

    await subscription_service.increment_usage(current_user.userId)
    await history_service.create_record({
        "userId": current_user.userId,
        "serviceId": SERVICE_ID,
        "serviceName": SERVICE_NAME,
        "fileName": document.filename,
        "fileSize": len(buffer),
        "format": "excel",
        "status": "success",
        "result": f"Extracted {result['total_rows']} rows across {result['total_pages']} pages",
        "processingTime": processing_time,
    })

    return {"success": True, "data": result}


@router.post("/extract-streaming")
async def production_sheets_extract_streaming(
    document: UploadFile = File(...),
    current_user: UserResponse = Depends(require_ps),
):
    """Extract and stream results as each page completes (JSONL)."""
    buffer = await document.read()
    _validate_file(document.filename, len(buffer))

    can_process, _sub, message = await subscription_service.can_process(current_user.userId)
    if not can_process:
        raise HTTPException(
            status_code=403,
            detail=f"Processing limit reached. {message}. Please upgrade your plan.",
        )

    start = time.time()

    async def stream_results():
        try:
            async for chunk in production_sheets_service.extract_streaming(buffer, document.filename):
                yield chunk
        except ValueError as e:
            yield json.dumps({"type": "error", "error": str(e)}).encode() + b"\n"
        except Exception as e:
            print(f"[ProductionSheetsRouter] Streaming extraction error: {e}")
            import traceback
            traceback.print_exc()
            yield json.dumps({"type": "error", "error": str(e)}).encode() + b"\n"
        finally:
            processing_time = int((time.time() - start) * 1000)
            await subscription_service.increment_usage(current_user.userId)
            await history_service.create_record({
                "userId": current_user.userId,
                "serviceId": SERVICE_ID,
                "serviceName": SERVICE_NAME,
                "fileName": document.filename,
                "fileSize": len(buffer),
                "format": "excel",
                "status": "success",
                "result": "Streaming extraction completed",
                "processingTime": processing_time,
            })

    return StreamingResponse(
        stream_results(),
        media_type="application/x-ndjson",
        headers={"X-Content-Type-Options": "nosniff"},
    )


# ── Excel export ─────────────────────────────────────────────────────────────

_HEADER_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
_TITLE_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="000000")
_CENTER = Alignment(horizontal="center", vertical="center")
_THIN = Side(style="thin", color="000000")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _safe_sheet_title(name: str, used: set) -> str:
    """Excel sheet titles: max 31 chars, unique, no invalid chars."""
    for ch in "[]:*?/\\":
        name = name.replace(ch, " ")
    name = name.strip()[:31] or "Sheet"
    base, n = name, 1
    while name in used:
        suffix = f" {n}"
        name = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def _write_sheet(ws, page: Dict[str, Any]):
    """Render one production page into the canonical Abhitex layout:
        Row 1: DATE : | <date> | ... | SHIFT : | <shift>
        Row 2: <merged, centered> <SHEET TITLE> across all 7 columns
        Row 3: green column headers (LOT NO ... KGS)
        Rows 4+: data
    """
    sheet_type = page.get("sheet_type", "") or ""
    meta = sheet_meta(sheet_type)
    title = page.get("sheet_title") or meta["title"]
    team_header = page.get("team_header") or meta["team_header"]
    date_value = page.get("date", "") or ""
    shift_value = (page.get("shift", "") or "").strip().upper()
    if shift_value not in ("DAY", "NIGHT"):
        shift_value = ""
    rows = page.get("rows", []) or []

    n_cols = len(COLUMNS)

    # Row 1: DATE label + value (cols 1-2), SHIFT label + value (cols 6-7).
    ws.cell(row=1, column=1, value="DATE :").font = _HEADER_FONT
    ws.cell(row=1, column=2, value=date_value)
    shift_label = ws.cell(row=1, column=n_cols - 1, value="SHIFT :")
    shift_label.font = _HEADER_FONT
    shift_label.alignment = _CENTER
    ws.cell(row=1, column=n_cols, value=shift_value).alignment = _CENTER

    # Row 2: centered title banner spanning all columns
    title_cell = ws.cell(row=2, column=1, value=title)
    title_cell.font = _HEADER_FONT
    title_cell.fill = _TITLE_FILL
    title_cell.alignment = _CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

    # Row 3: green column headers (TeamCode label varies by sheet)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        label = team_header if col_name == "TeamCode" else col_name
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER

    # Rows 4+: data
    for r_offset, row in enumerate(rows, start=4):
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r_offset, column=col_idx, value=str(row.get(col_name, "") or ""))
            cell.border = _BORDER

    # Column widths
    widths = {
        "LOT NO": 12, "STYLE CODE": 14, "IO NUMBER": 14, "ShadeCode": 12,
        "TeamCode": 12, "PCS": 10, "KGS": 10,
    }
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_name, 12)


@router.post("/export")
async def production_sheets_export(
    body: ExportRequest,
    current_user: UserResponse = Depends(require_ps),
):
    """Export extracted data preserving the source sheet layout, one worksheet
    per extracted page. The worksheet is named after the detected sheet type so
    LENGTH HEMING / LENGTH CUTING / FNS TO JOB / CROSS CUTING pages stay
    distinguishable.
    """
    if not body.pages:
        raise HTTPException(400, "pages payload is required.")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    used_titles: set = set()
    # Count pages per label so single-page sheets get a clean name.
    label_counts: Dict[str, int] = {}
    for page in body.pages:
        lbl = (page.get("sheet_label") or sheet_meta(page.get("sheet_type", "")).get("label", "Sheet"))
        label_counts[lbl] = label_counts.get(lbl, 0) + 1

    seen: Dict[str, int] = {}
    for idx, page in enumerate(body.pages, start=1):
        meta = sheet_meta(page.get("sheet_type", ""))
        label = page.get("sheet_label") or meta.get("label", "Sheet")
        if label_counts.get(label, 0) > 1:
            seen[label] = seen.get(label, 0) + 1
            sheet_name = f"{label} {seen[label]}"
        else:
            sheet_name = label
        ws = wb.create_sheet(title=_safe_sheet_title(sheet_name, used_titles))
        _write_sheet(ws, page)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{body.title}_{int(time.time())}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
