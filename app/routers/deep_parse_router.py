from fastapi import APIRouter, UploadFile, File, HTTPException, Depends
from fastapi.responses import Response
import time
import io
import openpyxl
from app.services.deep_parse.deep_parse_service import deep_parse_service
from app.utils.auth import get_current_user
from app.models.user import UserResponse

router = APIRouter()

@router.post("/extract")
async def extract_deep_parse(
    document: UploadFile = File(...),
    current_user: UserResponse = Depends(get_current_user)
):
    try:
        buffer = await document.read()
        result = await deep_parse_service.extract_multi_page(buffer, current_user.userId)
        return {
            "success": True,
            "data": result
        }
    except Exception as e:
        print(f"[DeepParseRouter] Extract Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/export")
async def export_deep_parse(
    data: dict,
    current_user: UserResponse = Depends(get_current_user)
):
    """
    Exports all validated records to Excel.
    Each page/record becomes one row in the spreadsheet.
    
    Expects data format:
    {
      "validated_records": [
        {
          "page_number": 1,
          "fields": {
            "supplier_name": {"value": "...", "edited_value": "..."},
            ...
          }
        }
      ]
    }
    """
    try:
        records = data.get("validated_records", [])
        if not records:
            raise HTTPException(status_code=400, detail="No records to export")

        # Create new Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Deep Parse Export"

        # Define standard fields in order with EXACT Excel column names
        field_to_header = {
            "document_id": "Document ID",
            "supplier_name": "Supplier Name",
            "gst_no": "GST No.",
            "invoice_no": "Invoice No.",
            "invoice_date": "Invoice Date",
            "buyer_name": "Buyer Name",
            "buyer_gst": "Buyer GST",
            "buyer_address": "Buyer Address",
            "challan_no": "Challan No.",
            "challan_date": "Challan Date",
            "gate_entry_no": "Gate Entry No.",
            "gate_entry_date": "Gate Entry Date",
            "po_number": "PO Number",
            "item_code": "Item Code",
            "item_description": "Item Description",
            "hsn_code": "HSN Code",
            "quantity": "Quantity",
            "unit_price": "Unit Price",
            "total_amount": "Total Amount",
            "cgst_amount": "CGST Amount",
            "sgst_amount": "SGST Amount",
            "igst_amount": "IGST Amount",
            "total_tax_amount": "Total Tax Amount",
            "grand_total": "Grand Total",
            "discount": "Discount",
        }

        standard_fields = list(field_to_header.keys())
        
        # Collect all unique field keys from all records
        all_field_keys = set()
        for record in records:
            fields = record.get("fields", {})
            all_field_keys.update(fields.keys())
        
        # Build ordered fields: standard fields first, then any extra fields
        ordered_fields = []
        for f in standard_fields:
            if f in all_field_keys:
                ordered_fields.append(f)
                all_field_keys.discard(f)
        ordered_fields.extend(sorted(all_field_keys))  # Add any remaining fields
        
        # Create headers row with proper display names
        headers = []
        for f in ordered_fields:
            if f in field_to_header:
                headers.append(field_to_header[f])
            else:
                headers.append(f.replace('_', ' ').title())
        ws.append(headers)

        # Add data rows - one row per page/record
        for record in records:
            fields = record.get("fields", {})
            
            row = []
            for field_name in ordered_fields:
                field_obj = fields.get(field_name, {})
                # Use edited_value if present (user modified), else original value
                val = field_obj.get("edited_value")
                if val is None:
                    val = field_obj.get("value", "N/A")
                if val is None or val == "":
                    val = "N/A"
                row.append(val)
            
            ws.append(row)

        # Style header row
        header_fill = openpyxl.styles.PatternFill(start_color="FBBF24", end_color="FBBF24", fill_type="solid")
        header_font = openpyxl.styles.Font(bold=True, color="000000")
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"DeepParse_{len(records)}_Records_{int(time.time())}.xlsx"
        return Response(
            content=buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}"
            }
        )
    except Exception as e:
        print(f"[DeepParseRouter] Export Error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
