"""
Lot History Cards Router — API endpoints for the Abhitex
"LOT HISTORY CARD — GREY FOLDING" extractor.

Endpoints:
  POST /api/lot-history-cards/extract  → Extract one or more cards (one per page)
  POST /api/lot-history-cards/export   → Export the extracted cards as Excel
"""

from __future__ import annotations

import io
import time
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel

from app.services.lot_history_cards.lot_history_cards_service import (
    HEADER_FIELDS,
    ROPE_COLUMNS,
    ROPE_KEYS,
    lot_history_cards_service,
)
from app.services.subscription_service import subscription_service
from app.services.history_service import history_service
from app.utils.auth import get_current_user
from app.models.user import UserResponse


router = APIRouter()

ALLOWED_EXTENSIONS = {
    ".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".webp",
}
MAX_FILE_SIZE_MB = 20


def _validate_file(filename: str, size: int):
    if not filename:
        raise HTTPException(400, "No filename provided.")
    ext = ("." + filename.rsplit(".", 1)[-1].lower()) if "." in filename else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(415, f"Unsupported file type '{ext}'.")
    if size > MAX_FILE_SIZE_MB * 1024 * 1024:
        raise HTTPException(413, f"File too large. Max: {MAX_FILE_SIZE_MB} MB.")


class ExportRequest(BaseModel):
    pages: List[Dict[str, Any]] = []
    title: str = "Lot_History_Cards_Export"


@router.post("/extract")
async def lot_history_cards_extract(
    document: UploadFile = File(...),
    current_user: UserResponse = Depends(get_current_user),
):
    buffer = await document.read()
    _validate_file(document.filename, len(buffer))

    can_process, _sub, message = await subscription_service.can_process(current_user.userId)
    if not can_process:
        raise HTTPException(
            status_code=403,
            detail=f"Processing limit reached. {message}. Please upgrade your plan.",
        )

    start = time.time()
    try:
        result = await lot_history_cards_service.extract(buffer, document.filename)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        print(f"[LotHistoryCardsRouter] Extraction error: {e}")
        raise HTTPException(500, str(e))

    processing_time = int((time.time() - start) * 1000)

    await subscription_service.increment_usage(current_user.userId)

    await history_service.create_record({
        "userId": current_user.userId,
        "serviceId": "lot-history-cards",
        "serviceName": "Lot History Cards Extraction",
        "fileName": document.filename,
        "fileSize": len(buffer),
        "format": "excel",
        "status": "success",
        "result": (
            f"Extracted {result['total_rolls']} rolls across "
            f"{result['total_cards']} card(s)"
        ),
        "processingTime": processing_time,
    })

    return {"success": True, "data": result}


def _flat_columns() -> List[str]:
    """Column order for the flattened, one-row-per-roll export sheet."""
    cols: List[str] = ["Card #"] + list(HEADER_FIELDS) + ["Rope"]
    cols += list(ROPE_COLUMNS)
    return cols


@router.post("/export")
async def lot_history_cards_export(
    body: ExportRequest,
    current_user: UserResponse = Depends(get_current_user),
):
    """Export extracted cards.

    Produces a single flat sheet with one row per roll. The header fields
    (I.O. No, Dye Lot No, Shade No, Quality M.No) repeat on each of a card's
    rope rows, alongside which Rope the roll belongs to and the roll values.
    This keeps thousands of cards queryable/filterable in one table.
    """
    if not body.pages:
        raise HTTPException(400, "pages payload is required.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lot History Cards"

    header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    header_font = Font(bold=True, color="000000")
    center = Alignment(horizontal="center", vertical="center")

    columns = _flat_columns()
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    out_row = 2
    for card_idx, page in enumerate(body.pages, start=1):
        header = page.get("header", {}) or {}
        wrote_any = False
        for rope_key in ROPE_KEYS:
            for roll in (page.get(rope_key, []) or []):
                values = [card_idx]
                values += [str(header.get(f, "") or "") for f in HEADER_FIELDS]
                values.append(rope_key)
                values += [str(roll.get(c, "") or "") for c in ROPE_COLUMNS]
                for col_idx, val in enumerate(values, start=1):
                    ws.cell(row=out_row, column=col_idx, value=val)
                out_row += 1
                wrote_any = True

        # A card with no rolls still gets one header-only row so it isn't lost.
        if not wrote_any:
            values = [card_idx]
            values += [str(header.get(f, "") or "") for f in HEADER_FIELDS]
            values.append("")
            values += ["" for _ in ROPE_COLUMNS]
            for col_idx, val in enumerate(values, start=1):
                ws.cell(row=out_row, column=col_idx, value=val)
            out_row += 1

    # Auto column widths
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if col_letter is None and hasattr(cell, "column_letter"):
                col_letter = cell.column_letter
            try:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 2, 24)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{body.title}_{int(time.time())}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
