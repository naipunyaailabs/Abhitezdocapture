"""
Export Engine for Register Extractor — produces Excel/CSV from extracted rows.
"""

from __future__ import annotations

import io
import time
from typing import Any, Dict, List

import openpyxl
from fastapi.responses import Response


def export_register_data(
    rows: List[Dict[str, Any]],
    headers: List[str],
    fmt: str = "excel",
    title: str = "Register_Export",
) -> Response:
    """Export rows to Excel (.xlsx). fmt is currently always 'excel'."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Register Export"

    # Header row
    ws.append(headers)

    header_fill = openpyxl.styles.PatternFill(
        start_color="6366F1", end_color="6366F1", fill_type="solid"
    )
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    # Data rows
    for row in rows:
        ws.append([str(row.get(h, "") or "") for h in headers])

    # Auto column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{title}_{int(time.time())}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
