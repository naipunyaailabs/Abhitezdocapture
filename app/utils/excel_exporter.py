import openpyxl
import io
import os
from typing import Dict, Any

class ExcelExporter:
    TEMPLATE_PATH = r"f:\OfficeFolders\DoCapture\GoodsReceipt-Template 1.xlsx"

    def export_to_template(self, data: Dict[str, Any], document_id: str = "") -> bytes:
        if not os.path.exists(self.TEMPLATE_PATH):
            raise FileNotFoundError(f"Template not found at {self.TEMPLATE_PATH}")

        wb = openpyxl.load_workbook(self.TEMPLATE_PATH)
        ws = wb.active

        # Get line items
        line_items = data.get("line_items", [])
        if not isinstance(line_items, list):
            line_items = []

        # Start from row 2 (row 1 is headers)
        start_row = 2
        
        # Mapping global fields
        supplier_name = data.get("supplier_name", "")
        gst_no = data.get("gst_no", "")
        invoice_no = data.get("invoice_no", "")
        invoice_date = data.get("invoice_date", "")
        challan_no = data.get("challan_no", "")
        challan_date = data.get("challan_date", "")
        gate_entry_no = data.get("gate_entry_no", "")
        gate_entry_date = data.get("gate_entry_date", "")
        po_number = data.get("po_number", "")

        for i, item in enumerate(line_items):
            current_row = start_row + i
            
            # Fill common fields
            ws.cell(row=current_row, column=1, value=supplier_name)
            ws.cell(row=current_row, column=2, value=gst_no)
            ws.cell(row=current_row, column=3, value=invoice_no)
            ws.cell(row=current_row, column=4, value=invoice_date)
            ws.cell(row=current_row, column=5, value=challan_no)
            ws.cell(row=current_row, column=6, value=challan_date)
            ws.cell(row=current_row, column=7, value=gate_entry_no)
            ws.cell(row=current_row, column=8, value=gate_entry_date)
            ws.cell(row=current_row, column=9, value=document_id)
            ws.cell(row=current_row, column=10, value=po_number)
            
            # Fill item specific fields
            ws.cell(row=current_row, column=11, value=item.get("item_code", ""))
            ws.cell(row=current_row, column=12, value=item.get("item_description", ""))
            ws.cell(row=current_row, column=13, value=item.get("lot_no", ""))
            ws.cell(row=current_row, column=14, value=item.get("shade_no", ""))
            ws.cell(row=current_row, column=15, value=item.get("quantity", ""))
            ws.cell(row=current_row, column=16, value=item.get("rate", ""))
            ws.cell(row=current_row, column=17, value=item.get("gst", ""))
            ws.cell(row=current_row, column=18, value=item.get("item_amount", ""))

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

excel_exporter = ExcelExporter()
